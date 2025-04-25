import csv
import hashlib
import requests
from io import BytesIO
from pathlib import Path
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import tkinter as tk
from tkinter import filedialog, messagebox


def capitalizar(texto):
    return ' '.join(p.capitalize() for p in texto.strip().lower().split())


def hash_gravatar(email):
    email = email.strip().lower()
    return hashlib.sha256(email.encode('utf-8')).hexdigest()


def descargar_imagen(gravatar_hash):
    url = f"http://www.gravatar.com/avatar/{gravatar_hash}.png?s=500&default=mp"
    response = requests.get(url)
    if response.status_code == 200:
        return BytesIO(response.content)
    else:
        print(f"No se pudo descargar la imagen para {gravatar_hash}")
        return None


def procesar_csv_a_excel(archivo_csv, archivo_excel):
    wb = Workbook()
    ws = wb.active
    ws.title = "Procesado"

    # Escribir cabeceras
    ws.append(["Nombre", "Apellidos", "Gravatar"])

    with open(archivo_csv, newline='', encoding='utf-8') as f:
        reader = csv.reader(f)
        next(reader)  # Saltar cabecera

        for i, fila in enumerate(reader, start=2):  # Empieza en fila 2 por la cabecera
            if len(fila) < 5:
                continue  # Saltar filas incompletas

            nombre = capitalizar(fila[2])
            apellidos = capitalizar(fila[3])
            email = fila[4].strip().lower()

            # Calcular hash y descargar imagen
            gravatar_hash = hash_gravatar(email)
            imagen_bytes = descargar_imagen(gravatar_hash)

            # Escribir nombre y apellidos
            ws.cell(row=i, column=1, value=nombre)
            ws.cell(row=i, column=2, value=apellidos)

            if imagen_bytes:
                imagen_path = Path(f"temp_img_{i}.png")
                with open(imagen_path, 'wb') as fimg:
                    fimg.write(imagen_bytes.getvalue())

                img = Image(str(imagen_path))
                img.width = 50
                img.height = 50
                ws.add_image(img, f"C{i}")

                imagen_path.unlink()  # Borrar imagen temporal

    wb.save(archivo_excel)


# ====================
# Interfaz Gráfica
# ====================
def seleccionar_archivo():
    archivo = filedialog.askopenfilename(
        title="Selecciona un archivo CSV",
        filetypes=[("CSV files", "*.csv")]
    )
    if not archivo:
        return

    salida = filedialog.asksaveasfilename(
        title="Guardar como archivo Excel",
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")]
    )
    if not salida:
        return

    try:
        procesar_csv_a_excel(archivo, salida)
        messagebox.showinfo("¡Listo!", f"Archivo generado correctamente:\n{salida}")
    except Exception as e:
        messagebox.showerror("Error", f"Ocurrió un error:\n{str(e)}")


# Crear la ventana principal
ventana = tk.Tk()
ventana.title("Conversor CSV a Excel con Gravatar")
ventana.geometry("400x200")

instrucciones = tk.Label(
    ventana,
    text="Haz clic en el botón para seleccionar tu archivo CSV.",
    wraplength=350,
    justify="center",
    font=("Arial", 12)
)
instrucciones.pack(pady=30)

boton = tk.Button(
    ventana,
    text="Seleccionar archivo CSV",
    command=seleccionar_archivo,
    font=("Arial", 12),
    bg="#4CAF50",
    fg="white",
    padx=20,
    pady=10
)
boton.pack()

ventana.mainloop()
